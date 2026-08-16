"""
Encrypt the blood collection lab data before it goes into the public
GitHub repo. The output file (blood_data.enc.json) contains ciphertext
only — without the password, the numbers are unreadable even though the
repo itself is public.

Usage:
    python encrypt_blood_data.py path/to/latest_lab_export.xlsx "your-strong-password"

Re-run this every time new lab results come in, then commit/upload the
resulting blood_data.enc.json to GitHub (blood_report.html never changes).

IMPORTANT — read this:
  - This protects the DATA FILE at rest (anyone browsing your public repo
    only sees ciphertext, not names/results).
  - It does NOT replace real access control. Anyone who has the password
    AND the page link can decrypt it in their browser. Treat the password
    like a shared secret: use something strong and unique (12+ random
    characters, not your project name or a birthday), and send it to your
    team through a separate channel (Signal/WhatsApp/verbally) — never
    commit it to GitHub or paste it in an email alongside the link.
  - If someone leaves the team or the password leaks, generate a new
    password and re-run this script — the old ciphertext becomes useless
    without the old password.
"""

import sys
import json
import base64
import os
import openpyxl
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

PBKDF2_ITERATIONS = 250_000  # must match the value in blood_report.html

NORMAL_RANGES = {
    # Fasting sugar: both hypo- and hyper-glycemia are clinically meaningful -> two-sided.
    "Fasting_Blood_Sugar": {"label": "Fasting Blood Sugar", "unit": "mmol/L", "low": 4.2, "high": 6.4},
    # Creatinine: elevated value is the clinically meaningful signal (kidney function) -> upper bound only.
    "Serum_Creatinine": {"label": "Serum Creatinine", "unit": "mg/dL", "low": None, "high": 0.9},
    "SGPT(ALT)": {"label": "ALT (SGPT)", "unit": "U/L", "low": None, "high": 34.0},
    # Total cholesterol: only "too high" is a clinical concern -> upper bound only.
    "SGPT_Total_Cholesterol": {"label": "Total Cholesterol", "unit": "mg/dL", "low": None, "high": 200},
    # HDL is protective — only "too low" is a clinical concern (high HDL is not flagged).
    "SGPT_HDL": {"label": "HDL", "unit": "mg/dL", "low": 35, "high": None},
    "SGPT_LDL": {"label": "LDL", "unit": "mg/dL", "low": None, "high": 200},
    # Triglycerides: only "too high" is a clinical concern -> upper bound only.
    "SGPT_Total_Triglycerides": {"label": "Triglycerides", "unit": "mg/dL", "low": None, "high": 200},
    "eGFR": {"label": "eGFR", "unit": "", "low": None, "high": None},
}

FIELD_MAP = {
    "HHID": "hhid",
    "Patient_Name": "name",
    "Age": "age",
    "Fasting_Blood_Sugar": "Fasting_Blood_Sugar",
    "Serum_Creatinine": "Serum_Creatinine",
    "SGPT(ALT)": "SGPT(ALT)",
    "SGPT_Total_Cholesterol": "SGPT_Total_Cholesterol",
    "SGPT_HDL": "SGPT_HDL",
    "SGPT_LDL": "SGPT_LDL",
    "SGPT_Total_Triglycerides": "SGPT_Total_Triglycerides",
    "       eGFR": "eGFR",
    "Blood_Collection_Date": "date",
}


def load_rows(xlsx_path: str, sheet_name: str = "Sheet1"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows


def to_records(rows):
    records = []
    for r in rows:
        rec = {}
        for src_col, dest_key in FIELD_MAP.items():
            v = r.get(src_col)
            if dest_key == "date" and hasattr(v, "strftime"):
                v = v.strftime("%Y-%m-%d")
            rec[dest_key] = v
        records.append(rec)
    return records


def encrypt(payload: dict, password: str) -> dict:
    salt = os.urandom(16)
    nonce = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=PBKDF2_ITERATIONS)
    key = kdf.derive(password.encode("utf-8"))
    aesgcm = AESGCM(key)
    plaintext = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    ciphertext = aesgcm.encrypt(nonce, plaintext, None)
    return {
        "salt": base64.b64encode(salt).decode(),
        "iv": base64.b64encode(nonce).decode(),
        "ciphertext": base64.b64encode(ciphertext).decode(),
        "iterations": PBKDF2_ITERATIONS,
    }


def main():
    if len(sys.argv) != 3:
        print('Usage: python encrypt_blood_data.py <path_to_xlsx> "<password>"')
        sys.exit(1)

    xlsx_path, password = sys.argv[1], sys.argv[2]
    if len(password) < 10:
        print("Warning: password is short. Use 12+ random characters for real protection.")

    rows = load_rows(xlsx_path)
    records = to_records(rows)
    payload = {"ranges": NORMAL_RANGES, "records": records}

    enc = encrypt(payload, password)
    with open("blood_data.enc.json", "w", encoding="utf-8") as f:
        json.dump(enc, f)

    print(f"Encrypted {len(records)} records -> blood_data.enc.json")
    print("Commit/upload blood_data.enc.json to GitHub. blood_report.html does not need to change.")


if __name__ == "__main__":
    main()
