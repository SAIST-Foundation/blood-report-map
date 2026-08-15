"""
Regenerate data.json for the blood report delivery map.

Usage:
    python update_map_data.py path/to/latest_kobo_export.xlsx

Run this every time your team exports a fresh copy from KoboToolbox,
then commit/push (or upload) the updated data.json to GitHub.
The map page (index.html) itself never needs to change.
"""

import sys
import json
import openpyxl

REQUIRED_COLS = {
    "household_id", "child_name", "blood_report_delivered", "received_by",
    "delivery_date", "landmark", "full_address",
    "_gps_location_latitude", "_gps_location_longitude",
}


def load_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Assumes the main delivery sheet is the first sheet in the export.
    # If your Kobo form structure changes, adjust the sheet name/index below.
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    missing = REQUIRED_COLS - set(header)
    if missing:
        raise ValueError(
            f"Expected columns not found in the sheet: {missing}\n"
            f"Check that the Kobo form/export hasn't changed its field names."
        )

    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    return rows


def to_map_data(rows):
    data = []
    for r in rows:
        lat = r.get("_gps_location_latitude")
        lon = r.get("_gps_location_longitude")
        if lat is None or lon is None:
            continue  # skip rows with no GPS captured yet
        ddate = r.get("delivery_date")
        ddate = ddate.strftime("%Y-%m-%d") if hasattr(ddate, "strftime") else (ddate or "")
        data.append({
            "hid": r.get("household_id"),
            "name": r.get("child_name"),
            "status": r.get("blood_report_delivered"),
            "date": ddate,
            "by": r.get("received_by") or "",
            "landmark": r.get("landmark") or "",
            "address": r.get("full_address") or "",
            "lat": lat,
            "lon": lon,
        })
    return data


def main():
    if len(sys.argv) != 2:
        print("Usage: python update_map_data.py <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    rows = load_rows(xlsx_path)
    data = to_map_data(rows)

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    delivered = sum(1 for d in data if d["status"] == "yes")
    print(f"data.json updated: {len(data)} mapped records ({delivered} delivered, {len(data) - delivered} not yet).")
    print("Now commit + push (or re-upload) data.json to GitHub.")


if __name__ == "__main__":
    main()
